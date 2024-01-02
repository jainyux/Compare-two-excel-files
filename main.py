import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def compare_and_highlight_changes(file_p, file_s):
    try:
        # Read both Excel files
        df_p = pd.read_excel(file_p, sheet_name='Sheet1', engine='openpyxl')
        df_s = pd.read_excel(file_s, sheet_name='Sheet1', engine='openpyxl')

        # Merge dataframes based on the first column (position 0)
        merged_df = pd.merge(df_p, df_s, left_on=df_p.columns[0], right_on=df_s.columns[0], how='outer', suffixes=('_p', '_s'))

        # Create a Pandas Excel writer using openpyxl
        with pd.ExcelWriter(file_p, engine='openpyxl') as writer:
            writer.book = load_workbook(file_p)

            # Access the openpyxl Excel writer object from the Pandas Excel writer
            writer.sheets = {ws.title: ws for ws in writer.book.worksheets}

            # Get the sheet and apply formatting
            sheet_name = 'Sheet1'
            sheet = writer.sheets[sheet_name]

            # Check if the sheet is hidden and make it visible
            if sheet.sheet_state == 'hidden':
                sheet.sheet_state = 'visible'

            # Iterate over rows
            for idx, row in merged_df.iterrows():
                if pd.notna(row['C_p']) and pd.notna(row['C_s']) and row['C_p'] != row['C_s']:
                    # Highlight the cell in column C in yellow if there's a difference
                    sheet[f'C{idx + 2}'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

            # Find new additions in column A and highlight them in green
            max_row = sheet.max_row
            for idx, value in enumerate(df_p.iloc[:, 0]):
                if pd.notna(value) and value not in df_s.iloc[:, 0].values:
                    # Highlight the entire row in green
                    for col in range(1, sheet.max_column + 1):
                        sheet.cell(row=max_row + idx + 1, column=col).fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

            # Save changes to the Excel file
            writer.save()

    except Exception as e:
        print(f"An error occurred: {e}")

# Replace 'p.xlsx' and 's.xlsx' with your actual file paths
compare_and_highlight_changes(r'C:\Users\mjainy\OneDrive - Digicorner\Desktop\Python\Compare\p.xlsx', r'C:\Users\mjainy\OneDrive - Digicorner\Desktop\Python\Compare\s.xlsx')
